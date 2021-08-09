from openpyxl import load_workbook

from constant import CITIES

def read_processes(filename):
    print('Parser: reading processes from excel file')
    workbook = load_workbook(filename=filename, read_only=True)
    sheet = workbook.active
    processes = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        process_id = row[0]
        court = row[1]
        if 'revisarlo en tribunal' in court.lower():
            process_id = process_id[:-1] + '1'
        city, city_value = parse_to_city(court)
        entity = parse_to_entity(court, city)
        process = {
            'id': process_id,
            'city': city_value,
            'entity': entity
        }
        processes.append(process)
    return processes 

def parse_to_city(court):
    for city, value in CITIES.items():
        if city in court.upper():
            return city, value['value']
    default_city = 'MEDELLIN'
    return default_city, CITIES[default_city]['value']

def parse_to_entity(court, city):
    entities = CITIES[city]['entities']
    for entity, value in entities.items():
        if entity in court.upper():
            return value
